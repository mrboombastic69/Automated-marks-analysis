import re
import io
import os
from openpyxl import load_workbook
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet,ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image
from reportlab.platypus import Spacer
from reportlab.lib.units import inch
class ResultAnalysis:
    def __init__(self, filename, start_row=8):
        self.workbook = load_workbook(filename=filename, data_only=True)
        self.sheet = self.workbook.active
        self.start_row = start_row
        self.subjects = {}
        self.sem = None
        self.year = None
        self.month = None
        # Define regex patterns for headings and subheadings
        self.heading_pattern = re.compile(r'^\d{2}[A-Za-z]{2,4}\d{2,3}$')
        self.subheading_pattern = re.compile(r'^(int|ext|tot)$', re.IGNORECASE)
    def extract_data(self):
        current_subject = None
        subheading_values = None
        for row in self.sheet.iter_rows(min_row=self.start_row, max_row=self.sheet.max_row):
            for cell in row:
                if cell.value is not None:
                    cell_value = str(cell.value).strip()
                    # Check if the cell value matches the heading pattern
                    if self.heading_pattern.match(cell_value):
                        current_subject = cell_value
                        self.subjects[current_subject] = {'int': [], 'ext': [], 'tot': []}
                        # Get the merged cell range
                        merged_range = None
                        for merged in self.sheet.merged_cells.ranges:
                            if cell.coordinate in merged:
                                merged_range = merged
                                break
                        if merged_range:
                            subheading_row = merged_range.min_row + 1
                            subheading_cols = range(merged_range.min_col, merged_range.max_col + 1)
                            # Extract the subheading values (int, ext, tot)
                            subheading_values = {}
                            for col in subheading_cols:
                                sub_cell = self.sheet.cell(row=subheading_row, column=col)
                                sub_cell_value = str(sub_cell.value).strip().lower()
                                if self.subheading_pattern.match(sub_cell_value):
                                    subheading_values[sub_cell_value] = col
                            # Iterate through the rows under the subheading to extract marks
                            for data_row in self.sheet.iter_rows(min_row=subheading_row + 1, max_row=self.sheet.max_row):
                                for subheading, col in subheading_values.items():
                                    data_cell = data_row[col - 1].value
                                    if isinstance(data_cell, (int, float)):
                                        self.subjects[current_subject][subheading].append(data_cell)

    def analyze(self, range_fcd=(75, 100), range_fc=(60, 74), range_sc=(35, 59), range_fail=(0, 34)):
        result = {}
        # Iterate through each subject
        for subject, sub_marks in self.subjects.items():
            n_fcd = 0
            n_fc = 0
            n_sc = 0
            n_fail = 0
            # Get the total marks
            tot_marks = sub_marks['tot']
            # Iterate through each total mark and classify
            for mark in tot_marks:
                if range_fcd[0] <= mark <= range_fcd[1]:
                    n_fcd += 1
                elif range_fc[0] <= mark <= range_fc[1]:
                    n_fc += 1
                elif range_sc[0] <= mark <= range_sc[1]:
                    n_sc += 1
                elif range_fail[0] <= mark <= range_fail[1]:
                    n_fail += 1
            # Store the result for the current subject
            result[subject] = {
                'n_fcd': n_fcd,
                'n_fc': n_fc,
                'n_sc': n_sc,
                'n_fail': n_fail
            }
        return result
    def extract_exam_info(self):
        # Initialize the dictionary to store the extracted info
        exam_info = {'year': None, 'sem': None, 'month': None}

        # Extract the data from the 7th row
        row = self.sheet[7]

        for cell in row:
            if cell.value is not None:
                cell_value = str(cell.value).strip()

                # Extract examination month and year
                if cell_value.startswith('Examination Month & Year:'):
                    parts = cell_value.split(':')
                    if len(parts) > 1:
                        month_year = parts[1].strip().split(' ')
                        if len(month_year) >= 2:
                            exam_info['month'] = month_year[0]
                            exam_info['year'] = month_year[1]

                # Extract semester
                elif cell_value.startswith('Sem:'):
                    parts = cell_value.split(':')
                    if len(parts) > 1:
                        exam_info['sem'] = parts[1].strip()

        return exam_info
    def create_pdf_report(self, result, filename='result_report.pdf'):
        # Create a PDF document
        pdf = SimpleDocTemplate(filename, pagesize=landscape(A4), rightMargin=36, leftMargin=36, topMargin=0, bottomMargin=0)
        story = []

        # Extract and save image from Excel sheet
        excel_image = None
        for image in self.sheet._images:
            excel_image = image
            break  # Assuming you want the first image

        if excel_image:
            # Extract image data from Excel image
            image_stream = io.BytesIO(excel_image._data())

            # Save image to file
            image_path = 'extracted_image.png'
            with open(image_path, 'wb') as f:
                f.write(image_stream.getvalue())

            # Add image to PDF
            pdf_image = Image(image_path, width=pdf.width - 72, height=100)  # Adjusted height
            story.append(pdf_image)
            story.append(Spacer(1, 12))

        # Include extracted examination information
        exam_info = self.extract_exam_info()
        exam_style = ParagraphStyle(name='ExamInfoStyle', fontName='Times-Bold', fontSize=16, spaceAfter=14)
        exam_paragraph = Paragraph(f"EXAMINATION MONTH & YEAR: {exam_info['month']} {exam_info['year']}<br/><br/>SEMESTER: {exam_info['sem']}", exam_style)
        story.append(exam_paragraph)
        story.append(Spacer(1, 24))
        # Create the table data
        table_data = [['Subject', 'FCD', 'FC', 'SC', 'Fail']]
        for subject, counts in result.items():
            row = [subject, counts['n_fcd'], counts['n_fc'], counts['n_sc'], counts['n_fail']]
            table_data.append(row)

        # Create the table
        table = Table(table_data, colWidths=[pdf.width/5]*5)
        table.setStyle(TableStyle([
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Times-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 18),  # Larger text for header
            ('FONTSIZE', (0, 1), (-1, -1), 16),  # Larger text for body
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),  # Increase cell height
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(table)
        story.append(Spacer(1, 24))

        # Signature
        signature_style = getSampleStyleSheet()['Normal']
        signature_style.fontName = 'Times-Bold'
        signature_style.fontSize = 20
        signature = Paragraph('Head of Department<br/><br/>Dr Ramakrishna M', signature_style)
        story.append(Spacer(1, 50))  # Spacer for positioning
        story.append(signature)

        pdf.build(story)

        # Delete the image file after the report is generated
        if os.path.exists(image_path):
            os.remove(image_path)

def result_analysis(infile, outfile_path):
    processor = ResultAnalysis(filename=infile)
    processor.extract_data()

    t = processor.extract_exam_info()
    fname = 'SEM-' + t.get('sem') + '-' + t.get('month') + '-' + t.get('year') + '.pdf'
    outfile = os.path.join(outfile_path, fname)

    result = processor.analyze()
    processor.create_pdf_report(result, outfile)

    return fname 